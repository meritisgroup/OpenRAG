import pandas as pd
import shutil
import json
import os
import fitz
import re
import unidecode
from pptx import Presentation
from openpyxl import load_workbook

def list_files(path_folder: str) -> list:
    if os.path.exists(path_folder):
        files = [file for file in os.listdir(path_folder) if os.path.isfile(os.path.join(path_folder, file))]
        return files
    else:
        return f"Le dossier {path_folder} n'existe pas."

def list_folders(path_folder: str) -> list:
    if os.path.exists(path_folder):
        folders = [folder for folder in os.listdir(path_folder) if os.path.isdir(os.path.join(path_folder, folder))]
        return folders
    else:
        return f"Le dossier {path_folder} n'existe pas."

def open_queries(path_queries: str) -> list[str]:
    with open(path_queries, 'r', encoding='utf-8') as f:
        content = f.readlines()
    for (k, line) in enumerate(content):
        content[k] = line.replace('\n', '')
    return content

def read_dictionary(file):
    with open(file, 'r') as f:
        content = f.read()
    blocs = content.strip().split('\n\n')
    dictionaries = [json.loads(bloc) for bloc in blocs]
    return dictionaries

def check_path(path_input: str, path_output: str) -> str:
    if path_input[-1] != '/':
        path_input += '/'
    if path_output[-1] != '/':
        path_output += '/'
    if 'raw_docs' not in path_input:
        try:
            former_path_input = path_input
            path_input += 'raw_docs/'
            os.mkdir(path_input)
            for file in list_files(former_path_input):
                if file != 'queries.txt' and file != 'answers.txt':
                    shutil.move(former_path_input + file, path_input)
        except Exception as e:
            pass
    try:
        os.mkdir(path_input.replace('raw_docs', 'txt_docs'))
    except Exception as e:
        pass
    try:
        os.mkdir(path_input.replace('raw_docs', 'entities_docs'))
    except Exception as e:
        pass
    return (path_input, path_output)

def save_answers(queries: list[str], answers: list[str], path_folder: str, name: str) -> None:
    df = pd.DataFrame(columns=['Query', 'Answer'])
    clock = 0
    for (query, answer) in zip(queries, answers):
        df.loc[clock] = {'Query': query, 'Answer': answer}
        clock += 1
    if path_folder[-1] != '/':
        path_folder += '/'
    df.to_excel(path_folder + name + '.xlsx')

def open_real_answers(path_txt_answers: str) -> list[str]:
    return open_queries(path_txt_answers)

def open_proposed_answers(path_df_answers: str) -> list[str]:
    proposed_answers = []
    df = pd.read_excel(path_df_answers).drop(['Unnamed: 0'], axis=1)
    for k in range(df.shape[0]):
        if list(df['Query'])[k] != 'Mean':
            proposed_answers.append(list(df['Answer'])[k])
    return proposed_answers

def replace_extension(path: str, replace_value: str) -> str:
    if '.' not in path:
        return path
    else:
        clock = -1
        while path[clock] != '.':
            clock -= 1
        return path[:clock] + replace_value

def get_extension(text: str) -> str:
    if type(text) is not type('str') or '.' not in text:
        return None
    else:
        return '.' + text.split('.')[-1]

def count_page(path_fichier: str) -> int:
    if get_extension(path_fichier) == '.pdf':
        document = fitz.open(path_fichier)
        nombre_pages = document.page_count
        document.close()
        return nombre_pages
    elif get_extension(path_fichier) == '.pptx' or get_extension(path_fichier) == '.ppt':
        presentation = Presentation(path_fichier)
        nombre_diapos = len(presentation.slides)
        return nombre_diapos
    elif get_extension(path_fichier) == '.xlsx':
        classeur = load_workbook(path_fichier, read_only=True)
        nombre_feuilles = len(classeur.sheetnames)
        classeur.close()
        return nombre_feuilles
    else:
        return 0

def clean_text(text):
    text = unidecode.unidecode(text)
    return re.sub('[^\\w\\s]', '', text)